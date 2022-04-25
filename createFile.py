import openpyxl
import csv

wb = openpyxl.load_workbook("book_data.xlsx")
sh = wb['book data']
row = sh.max_row
column=sh.max_column
bookInfo = []
rating = []
isbn = sh.cell(2,3).value

for i in range(2,3):
    line= []
    ratingLine = []
    for j in range(1,column+1):
        if(j==2 or j==3 or j==4):
            ratingLine.append(sh.cell(i,j).value)
        elif(j==10):
            curstr = sh.cell(i,j).value.replace("\'","")
            curstr = curstr.replace("[","")
            curstr = curstr.replace("]","")
            line.append(str(curstr))
        else:
            line.append(sh.cell(i,j).value)
    rating.append(ratingLine)
    bookInfo.append(line)


for i in range(3,row+1):
    line= []
    ratingLine = []
    bid = ""
    for j in range(1,column+1):
        if(j==2 or j==3 or j==4):
            ratingLine.append(sh.cell(i,j).value)

        if(j==10):
            if(sh.cell(i,j).value==9):
                line.append("Fiction")
            else:
                curstr = str(sh.cell(i,j).value)
                curstr = curstr.replace("\'","")
                curstr = curstr.replace("[","")
                curstr = curstr.replace("]","")
                line.append(str(curstr))
        elif(j!=4 and j!=2 and j!=1):
            line.append(sh.cell(i,j).value)

        if(j==3):
            bid = sh.cell(i,j).value
    rating.append(ratingLine)
    if(bid!=isbn):
        isbn=bid
        bookInfo.append(line)

Heads= ['book_id', 'book_title', 'book_author', 'year_of_publication','publisher','img_l','Category']
with open('book_info.csv','w+', newline='') as f:
    write = csv.writer(f)
    write.writerow(Heads)
    write.writerows(bookInfo)

Heads = ['user_id', 'book_id', 'rating']
with open('rating.csv','w+',newline='') as f:
    write = csv.writer(f)
    write.writerow(Heads)
    write.writerows(rating)